"""연락처 라우터 - CRUD + Excel"""
from typing import List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.user import User
from app.models.contact import Contact
from app.schemas.contact import ContactCreate, ContactUpdate, ContactResponse
from app.middleware.auth import get_current_user
import io

router = APIRouter(prefix="/contacts", tags=["연락처"])


@router.get("", response_model=List[ContactResponse])
async def list_contacts(
    category: str = None, search: str = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    query = select(Contact).where(Contact.user_id == user.id)
    if category and category != "all":
        query = query.where(Contact.category == category)
    query = query.order_by(Contact.name)

    result = await db.execute(query)
    contacts = list(result.scalars().all())

    if search:
        s = search.lower()
        contacts = [c for c in contacts
                    if s in c.name.lower() or s in c.company.lower()
                    or s in c.memo.lower() or s in c.phone]
    return contacts


@router.post("", response_model=ContactResponse)
async def create_contact(
    req: ContactCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    contact = Contact(user_id=user.id, **req.model_dump())
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact


@router.put("/{contact_id}", response_model=ContactResponse)
async def update_contact(
    contact_id: int, req: ContactUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Contact).where(Contact.id == contact_id, Contact.user_id == user.id)
    )
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(404, "연락처를 찾을 수 없습니다")

    for key, val in req.model_dump(exclude_none=True).items():
        setattr(contact, key, val)
    await db.commit()
    await db.refresh(contact)
    return contact


@router.delete("/{contact_id}")
async def delete_contact(
    contact_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Contact).where(Contact.id == contact_id, Contact.user_id == user.id)
    )
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(404, "연락처를 찾을 수 없습니다")
    await db.delete(contact)
    await db.commit()
    return {"message": "삭제되었습니다"}


@router.post("/import")
async def import_contacts(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """엑셀 파일 업로드로 연락처 일괄 등록"""
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    col_map = {}
    mapping = {"이름": "name", "카테고리": "category", "전화번호": "phone",
               "회사": "company", "직급": "position", "메모": "memo",
               "생일": "birthday", "기념일": "anniversary"}
    for i, h in enumerate(headers):
        if h in mapping:
            col_map[mapping[h]] = i

    if "name" not in col_map:
        raise HTTPException(400, "'이름' 컬럼이 필요합니다")

    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[col_map["name"]] or "").strip()
        if not name:
            continue
        data = {"name": name}
        for field, idx in col_map.items():
            if field != "name" and idx < len(row):
                data[field] = str(row[idx] or "").strip()

        contact = Contact(user_id=user.id, **data)
        db.add(contact)
        added += 1

    await db.commit()
    wb.close()
    return {"added": added}


@router.get("/export")
async def export_contacts(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """엑셀 파일로 내보내기"""
    import openpyxl
    result = await db.execute(
        select(Contact).where(Contact.user_id == user.id).order_by(Contact.name)
    )
    contacts = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "연락처"
    headers = ["이름", "카테고리", "전화번호", "회사", "직급", "메모", "생일", "기념일"]
    ws.append(headers)

    for c in contacts:
        ws.append([c.name, c.category, c.phone, c.company, c.position,
                   c.memo, c.birthday, c.anniversary])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contacts.xlsx"}
    )
